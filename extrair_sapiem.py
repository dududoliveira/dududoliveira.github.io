#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extrair_sapiem.py — converte uma planilha de modalidades do SAPIEM e os PDFs das
normas municipais em um arquivo .json consumido pelo Painel_SAPIEM.html.

Uso basico:
    python extrair_sapiem.py planilha.xlsx --aba Cachoeirinha --id cachoeirinha \
        --norma "LOM=emenda_lom.pdf" --norma "LCM 83/2022=lei_complementar_83.pdf" \
        --saida municipios/cachoeirinha.json

O que faz:
  1. le a aba indicada da planilha e normaliza as 64 (ou N) modalidades;
  2. captura os comentarios de celula — costumam guardar as tabelas de progressao;
  3. faz o parser da coluna EMBASAMENTO_LEGAL e monta o indice de dispositivos;
  4. extrai o texto de cada artigo dos PDFs informados e casa com o indice;
  5. roda as conferencias automaticas e grava relatorio_<id>.md;
  6. grava o .json do municipio.

O que NAO faz: julgar. As divergencias entre a planilha e a lei que dependem de
leitura ficam no relatorio como pendencia para conferencia humana.

Dependencias: openpyxl (obrigatoria), pdfplumber (opcional — sem ela o script
tenta o binario pdftotext; sem os dois, o campo de textos sai vazio).
"""

import argparse
import json
import re
import subprocess
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")

from openpyxl.utils import get_column_letter

DIAS_ANO = 365
DIAS_MES = 30

# colunas de requisito da planilha -> chave interna usada pelo painel
REQUISITOS = {
    "IDADE_ATUAL**": "idade",
    "TEMPO_CONTRIBUICAO_ATUAL**": "contribuicao",
    "TEMPO_SERV_PUBLICO_ATUAL**": "servicoPublico",
    "TEMPO_CARREIRA_ATUAL": "carreira",
    "TEMPO_MAGISTERIO_ATUAL**": "magisterio",
    "TEMPO_CARGO_ATUAL**": "cargo",
    "SOMA_IDADE_CONTRIB_ATUAL**": "pontos",
}


# --------------------------------------------------------------------------- #
# utilidades
# --------------------------------------------------------------------------- #
def limpar(v):
    if v is None:
        return ""
    s = str(v).strip().strip("'").strip('"').strip()
    return re.sub(r"\s+", " ", s)


def sem_acento(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", (s or "").lower())
        if unicodedata.category(c) != "Mn"
    )


def achatar_paragrafos(texto):
    """Junta as quebras de linha internas de cada paragrafo, preservando os
    paragrafos (separados por linha em branco)."""
    blocos = re.split(r"\n\s*\n", texto)
    blocos = [re.sub(r"\s*\n\s*", " ", b).strip() for b in blocos]
    return "\n\n".join(b for b in blocos if b)


# --------------------------------------------------------------------------- #
# 1-3. planilha
# --------------------------------------------------------------------------- #
def ler_planilha(caminho, nome_aba):
    wb_val = openpyxl.load_workbook(caminho, data_only=True)
    wb_com = openpyxl.load_workbook(caminho)

    if nome_aba is None:
        nome_aba = wb_val.sheetnames[0]
    if nome_aba not in wb_val.sheetnames:
        sys.exit(
            f"ERRO: aba '{nome_aba}' nao existe. Abas disponiveis: "
            + ", ".join(wb_val.sheetnames)
        )

    ws, ws_com = wb_val[nome_aba], wb_com[nome_aba]
    linhas = list(ws.iter_rows(values_only=True))
    if not linhas:
        sys.exit(f"ERRO: aba '{nome_aba}' esta vazia.")

    cab = [str(h).strip() if h is not None else "" for h in linhas[0]]
    letra_para_campo = {get_column_letter(i + 1): h for i, h in enumerate(cab)}

    # comentarios de celula
    comentarios, legenda = {}, []
    for linha in ws_com.iter_rows():
        for cel in linha:
            if not cel.comment:
                continue
            txt = re.sub(r"\s*\n\s*", "\n", cel.comment.text.strip())
            autor = ""
            m = re.match(r"^([A-Za-zÀ-ÿ. ]{3,40}):\s*", txt)
            if m:
                autor = m.group(1).strip()
                txt = txt[m.end():].strip()
            campo = letra_para_campo.get(cel.column_letter, cel.column_letter)
            item = {"campo": campo, "autor": autor, "texto": txt}
            if cel.row == 1:
                legenda.append(item)
            else:
                comentarios.setdefault(cel.row, []).append(item)

    return cab, linhas, comentarios, legenda, nome_aba


def parse_refs(embasamento, siglas):
    """Extrai [{norma, artigo}] de 'art. 92-A da LOM c/c arts. 12 e 31 da LCM 83/2022'.

    Estrategia: localiza os marcadores de norma no texto e atribui a cada
    marcador os numeros de artigo que aparecem no trecho anterior a ele.
    """
    if not embasamento or not siglas:
        return []
    padrao = "|".join(re.escape(s) for s in sorted(siglas, key=len, reverse=True))
    marcas = [
        (m.start(), m.end(), m.group(0))
        for m in re.finditer(rf"(?:d[aeo]s?\s+)?({padrao})", embasamento)
    ]
    refs, vistos, anterior = [], set(), 0
    for ini, fim, bruto in marcas:
        norma = next(s for s in sorted(siglas, key=len, reverse=True) if s in bruto)
        trecho = embasamento[anterior:ini]
        anterior = fim
        trecho = re.sub(r"§+\s*\d+[ºo]?(?:-[A-Z])?", "", trecho)   # tira paragrafos
        trecho = re.sub(r"\b\d{2}/\d{2}/\d{4}\b", "", trecho)      # tira datas
        for m in re.finditer(r"\b(\d{1,3}(?:-[A-Z])?)\b", trecho):
            chave = (norma, m.group(1))
            if chave not in vistos:
                vistos.add(chave)
                refs.append({"norma": norma, "artigo": m.group(1)})
    return refs


def montar_modalidades(cab, linhas, comentarios, siglas, marcos):
    def pega(linha, campo):
        return limpar(linha[cab.index(campo)]) if campo in cab else ""

    col_2003 = "CARGO_EFETIVO_2003"
    col_2022 = next((c for c in cab if c.startswith("CARGO_EFETIVO_até")), None)
    col_1998 = "CARGO_EFETIVO_1998"

    saida = []
    for n, linha in enumerate(linhas[1:], 2):
        if not any(linha):
            continue
        nome = pega(linha, "DS_MODALIDADE_BENEFICIO")
        if not nome:
            continue
        nn = sem_acento(nome)

        req, progressivo = {}, False
        for col, chave in REQUISITOS.items():
            v = pega(linha, col)
            if not v:
                continue
            if chave == "pontos" and v in ("S", "N"):
                progressivo = v == "S"
                continue
            try:
                req[chave] = int(float(v))
            except ValueError:
                pass

        f03 = pega(linha, col_2003)
        f22 = pega(linha, col_2022) if col_2022 else ""
        f98 = pega(linha, col_1998)

        marco = ""
        for rotulo, cfg in marcos.items():
            if sem_acento(cfg["texto"]) in nn:
                marco = rotulo
                break
        if not marco:
            if f03 == "S":
                marco = "2003"
            elif f22 == "S":
                marco = "2022"

        grupos = []
        for chave, termo in [
            ("incapacidade", "incapacidade permanente"), ("professor", "professor"),
            ("pcd", "deficiencia"), ("nocivos", "agentes nocivos"),
            ("compulsoria", "compulsoria"),
        ]:
            if termo in nn:
                grupos.append(chave)
        if not grupos:
            grupos.append("comum")

        if "pontuacao" in nn:
            regra = "pontuacao"
        elif "pedagio" in nn:
            regra = "pedagio"
        elif "regra de transicao" in nn:
            regra = "transicao"
        else:
            regra = "permanente"

        grau = ""
        for g in ("grave", "moderado", "leve"):
            if "grau " + g in nn:
                grau = g
        if "deficiencia por idade" in nn:
            grau = "idade"

        achados = []
        for rotulo, cfg in marcos.items():
            if sem_acento(cfg["texto"]) in nn:
                coluna = col_2003 if rotulo == "2003" else col_2022
                if coluna and pega(linha, coluna) != "S":
                    achados.append(
                        f"Nome indica {cfg['rotulo']}, mas a coluna {coluna} esta vazia."
                    )
        if pega(linha, "ORDEM_EXIBICAO") == "#REF!":
            achados.append("ORDEM_EXIBICAO com erro de referencia (#REF!) na planilha.")

        saida.append({
            "linha": n,
            "nome": nome,
            "especie": pega(linha, "ID_ESPECIE_BENEFICIO"),
            "baseLegal": pega(linha, "EMBASAMENTO_LEGAL").rstrip('"').strip(),
            "tituloAto": pega(linha, "TITULO_ATO"),
            "descricaoAto": pega(linha, "DESCRICAO_ATO"),
            "sexo": pega(linha, "SEXO"),
            "voluntaria": pega(linha, "VOLUNTARIA"),
            "integProporc": pega(linha, "INTEG_PROPORC"),
            "media": pega(linha, "MEDIA"),
            "paridade": pega(linha, "PARIDADE_INATIVACAO"),
            "redutor": pega(linha, "REDUTOR"),
            "marco": marco,
            "grupos": grupos,
            "regra": regra,
            "grau": grau,
            "req": req,
            "pontosProgressivo": progressivo,
            "calculoMedia": pega(linha, "Cálculo da média"),
            "calculoBeneficio": pega(linha, "Cálculo do benefício"),
            "baseRMI": pega(linha, "Base RMI"),
            "pctAnoAcima20": pega(linha, "percentual_ano_acima_20_anos"),
            "pctMaiores": pega(linha, "PCT_MAIORES_CONTRIBUICOES"),
            "propValRef": pega(linha, "PROP_VAL_REFERENCIA"),
            "invalidez": pega(linha, "INVALIDEZ****"),
            "acidente": pega(linha, "ACIDENTE_SERVICO"),
            "molestia": pega(linha, "MOLESTIA_PROFISSIONAL"),
            "doencaLei": pega(linha, "DOENCA_ESPEC_LEI"),
            "doencaComum": pega(linha, "DOENCA_COMUM"),
            "f98": f98, "f03": f03, "f22": f22,
            "refs": parse_refs(pega(linha, "EMBASAMENTO_LEGAL"), siglas),
            "notas": comentarios.get(n, []),
            "achados": achados,
        })
    return saida


# --------------------------------------------------------------------------- #
# 4. PDFs das normas
# --------------------------------------------------------------------------- #
def texto_da_norma(caminho):
    """Retorna (texto, metodo). Texto vazio = PDF sem camada de texto.

    Aceita .pdf, .txt e .md — util quando o PDF e digitalizado e o texto
    precisou ser transcrito a mao.
    """
    if Path(caminho).suffix.lower() in (".txt", ".md"):
        return Path(caminho).read_text(encoding="utf-8"), "arquivo de texto"
    try:
        import pdfplumber
        with pdfplumber.open(caminho) as pdf:
            paginas = [p.extract_text() or "" for p in pdf.pages]
        txt = "\n".join(paginas)
        if len(txt.strip()) > 200:
            return txt, "pdfplumber"
    except ImportError:
        pass
    except Exception as e:
        print(f"  aviso: pdfplumber falhou em {caminho}: {e}")
    try:
        r = subprocess.run(["pdftotext", "-layout", str(caminho), "-"],
                           capture_output=True, text=True, timeout=120)
        if len(r.stdout.strip()) > 200:
            return r.stdout, "pdftotext"
    except Exception:
        pass
    return "", "nenhum"


def separar_artigos(texto):
    """Fatia o texto da norma em {numero_do_artigo: corpo}."""
    t = texto.replace("\x0c", "\n")
    limpas = []
    for linha in t.split("\n"):
        s = linha.strip()
        if re.fullmatch(r"\d{1,3}\s*/\s*\d{1,3}", s):        # "12/28"
            continue
        if s.lower().startswith("leismunicipais.com.br"):     # rodape
            continue
        limpas.append(linha.rstrip())
    t = "\n".join(limpas)
    t = re.sub(r"(Art\.\s*\d+(?:-[A-Z])?[ºo]?\.?)(?=[A-ZÀ-Ú])", r"\1 ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)

    marcas = list(re.finditer(r"(?:^|\n)Art\.?\s*(\d{1,3}(?:\s*-\s*[A-Z])?)[ºo]?\.?\s", t))
    artigos = {}
    for i, m in enumerate(marcas):
        num = re.sub(r"\s", "", m.group(1)).upper()
        ini = m.start() + (1 if t[m.start()] == "\n" else 0)
        fim = marcas[i + 1].start() if i + 1 < len(marcas) else len(t)
        corpo = t[ini:fim].strip()
        corpo = re.sub(r"\n\n(?:TÍTULO|CAPÍTULO|Seção|SEÇÃO)[\s\S]*$", "", corpo).strip()
        corpo = achatar_paragrafos(corpo)
        # a versao mais longa costuma ser a do corpo da lei, nao a de uma remissao
        if num not in artigos or len(corpo) > len(artigos[num]):
            artigos[num] = corpo
    return artigos


# --------------------------------------------------------------------------- #
# 5. conferencias automaticas
# --------------------------------------------------------------------------- #
def conferir(modalidades, textos, siglas):
    relatorio = {"internas": [], "textos": [], "cobertura": []}

    for m in modalidades:
        for a in m["achados"]:
            if not a.startswith("ORDEM_EXIBICAO"):
                relatorio["internas"].append(f"Linha {m['linha']} — {m['nome']}: {a}")
        if not m["refs"]:
            relatorio["internas"].append(
                f"Linha {m['linha']} — {m['nome']}: embasamento legal nao reconhecido "
                f"({m['baseLegal'][:80]!r}). Confira as siglas informadas em --norma."
            )
        if m["regra"] == "pontuacao" and not m["pontosProgressivo"]:
            relatorio["internas"].append(
                f"Linha {m['linha']} — {m['nome']}: regra de pontuacao sem a marca 'S' "
                f"em SOMA_IDADE_CONTRIB_ATUAL."
            )

    ref_erro = sum(1 for m in modalidades if any(a.startswith("ORDEM_EXIBICAO") for a in m["achados"]))
    if ref_erro:
        relatorio["internas"].append(
            f"ORDEM_EXIBICAO com #REF! em {ref_erro} de {len(modalidades)} linhas."
        )

    citados = {(r["norma"], r["artigo"]) for m in modalidades for r in m["refs"]}
    for norma, artigo in sorted(citados):
        if not textos.get(f"{norma}|{artigo}", "").strip():
            relatorio["textos"].append(f"{norma}, art. {artigo}: sem texto extraido.")

    for sigla in siglas:
        n = sum(1 for k in textos if k.startswith(sigla + "|"))
        relatorio["cobertura"].append(f"{sigla}: {n} dispositivos com texto.")

    return relatorio


def escrever_relatorio(destino, ident, municipio, modalidades, textos, siglas, rel):
    citados = sorted({(r["norma"], r["artigo"]) for m in modalidades for r in m["refs"]},
                     key=lambda x: (x[0], len(x[1]), x[1]))
    L = [f"# Relatorio de extracao — {municipio}", ""]
    L.append(f"- Modalidades processadas: **{len(modalidades)}**")
    L.append(f"- Dispositivos citados: **{len(citados)}**")
    L.append(f"- Normas: {', '.join(siglas) if siglas else '(nenhuma informada)'}")
    L += ["", "## Cobertura dos textos legais", ""]
    L += [f"- {x}" for x in rel["cobertura"]] or ["- (nenhuma norma informada)"]
    if rel["textos"]:
        L += ["", "### Dispositivos citados sem texto", ""]
        L += [f"- {x}" for x in rel["textos"]]
    L += ["", "## Conferencias automaticas", ""]
    L += [f"- {x}" for x in rel["internas"]] or ["- Nenhuma inconsistencia interna detectada."]

    L += ["", "## Para conferencia humana", "",
          "Compare os valores abaixo com o texto de cada dispositivo. O script nao",
          "julga se o parametro da planilha corresponde ao que a lei exige — essa",
          "leitura e sua. Atencao especial a: faixas de uma mesma regra que a planilha",
          "possa nao ter parametrizado; periodo adicional (pedagio), que costuma nao ter",
          "coluna propria; e datas-marco de ingresso.", ""]
    L += ["| Linha | Modalidade | Sexo | Requisitos (anos) | Base legal |",
          "|---|---|---|---|---|"]
    for m in modalidades:
        reqs = []
        for chave, dias in m["req"].items():
            anos = dias / DIAS_ANO
            reqs.append(f"{chave} {anos:g}")
        if m["pontosProgressivo"]:
            reqs.append("pontos progressivos")
        L.append(
            f"| {m['linha']} | {m['nome'][:70]} | {m['sexo'] or '—'} | "
            f"{'; '.join(reqs) or '—'} | {m['baseLegal'][:70]} |"
        )
    Path(destino).write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------- #
# principal
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Gera o .json de um municipio para o Painel SAPIEM.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("planilha", help="arquivo .xlsx do SAPIEM")
    ap.add_argument("--aba", help="nome da aba (padrao: a primeira)")
    ap.add_argument("--id", help="identificador curto, ex.: cachoeirinha")
    ap.add_argument("--municipio", help="nome exibido (padrao: nome da aba)")
    ap.add_argument("--norma", action="append", default=[], metavar="SIGLA=arquivo.pdf",
                    help='ex.: --norma "LCM 83/2022=lc83.pdf" — aceita .pdf, .txt ou .md (repetivel)')
    ap.add_argument("--textos", metavar="arquivo.json",
                    help="json {'SIGLA|artigo': 'texto'} para mesclar por cima do "
                         "que foi extraido — util para transcricoes manuais")
    ap.add_argument("--versao-planilha", default="", help="data da versao, ex.: 17/01/2023")
    ap.add_argument("--saida", help="caminho do .json (padrao: <id>.json)")
    args = ap.parse_args()

    ident = args.id or Path(args.planilha).stem.lower()
    ident = re.sub(r"[^a-z0-9-]+", "-", sem_acento(ident)).strip("-")

    print(f"→ lendo planilha {args.planilha}")
    cab, linhas, comentarios, legenda, aba = ler_planilha(args.planilha, args.aba)
    municipio = args.municipio or aba
    print(f"  aba '{aba}', {len(cab)} colunas, {len(linhas) - 1} linhas de dados")

    # normas
    normas, textos = [], {}
    for spec in args.norma:
        if "=" not in spec:
            sys.exit(f"ERRO: --norma esperava SIGLA=arquivo.pdf, recebeu {spec!r}")
        sigla, arquivo = spec.split("=", 1)
        sigla, arquivo = sigla.strip(), arquivo.strip()
        if not Path(arquivo).exists():
            sys.exit(f"ERRO: arquivo nao encontrado: {arquivo}")
        print(f"→ extraindo {sigla} de {arquivo}")
        bruto, metodo = texto_da_norma(arquivo)
        if not bruto:
            print(f"  ATENCAO: PDF sem camada de texto (provavelmente digitalizado). "
                  f"Os artigos de {sigla} ficarao vazios — cole o texto pelo painel, "
                  f"ou passe um .txt com o conteudo transcrito.")
            normas.append({"sigla": sigla, "arquivo": Path(arquivo).name,
                           "extracao": "falhou"})
            continue
        arts = separar_artigos(bruto)
        for num, corpo in arts.items():
            textos[f"{sigla}|{num}"] = corpo
        print(f"  {len(arts)} artigos ({metodo})")
        normas.append({"sigla": sigla, "arquivo": Path(arquivo).name, "extracao": metodo})

    if args.textos:
        extra = json.loads(Path(args.textos).read_text(encoding="utf-8"))
        extra = extra.get("textos", extra)
        textos.update({k: v for k, v in extra.items() if str(v).strip()})
        print(f"→ mesclados {len(extra)} dispositivos de {args.textos}")

    siglas = [n["sigla"] for n in normas]

    marcos = {
        "2003": {"texto": "ate 31/12/2003", "rotulo": "ingresso ate 31/12/2003"},
        "2022": {"texto": "ate 20/06/2022", "rotulo": "ingresso ate 20/06/2022"},
    }

    print("→ montando modalidades")
    modalidades = montar_modalidades(cab, linhas, comentarios, siglas, marcos)
    print(f"  {len(modalidades)} modalidades")

    # so os dispositivos citados entram no json (mantem o arquivo enxuto)
    citados = {f"{r['norma']}|{r['artigo']}" for m in modalidades for r in m["refs"]}
    textos_usados = {k: v for k, v in textos.items() if k in citados}
    print(f"  {len(citados)} dispositivos citados, {len(textos_usados)} com texto")

    print("→ conferindo")
    rel = conferir(modalidades, textos, siglas)

    dados = {
        "schema": "sapiem-municipio/1",
        "id": ident,
        "municipio": municipio,
        "fonte": Path(args.planilha).name,
        "versaoPlanilha": args.versao_planilha,
        "vigenciaConferidaEm": "",
        "normas": normas,
        "progressao": {},      # preencher conforme a lei do municipio
        "correlatos": [],      # dispositivos de consulta nao citados nas modalidades
        "achados": [],         # cruzamento planilha x lei, escrito por quem analisa
        "legenda": legenda,
        "modalidades": modalidades,
        "textos": textos_usados,
    }

    saida = Path(args.saida or f"{ident}.json")
    saida.parent.mkdir(parents=True, exist_ok=True)
    saida.write_text(json.dumps(dados, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"→ gravado {saida} ({saida.stat().st_size // 1024} KB)")

    rel_path = saida.parent / f"relatorio_{ident}.md"
    escrever_relatorio(rel_path, ident, municipio, modalidades, textos, siglas, rel)
    print(f"→ gravado {rel_path}")

    if rel["internas"] or rel["textos"]:
        print("\nPendencias (detalhe no relatorio):")
        for x in (rel["internas"] + rel["textos"])[:10]:
            print(f"  - {x}")
        extra = len(rel["internas"]) + len(rel["textos"]) - 10
        if extra > 0:
            print(f"  ... e mais {extra}.")


if __name__ == "__main__":
    main()
